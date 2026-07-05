from collections import defaultdict
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.entities import Completion, Punishment, Room, Student, Task
from app.services.analytics import BONUS_TASK_NAME, build_multi_room_analytics, build_room_analytics
from app.services.scoring import build_bonus_maps, build_leaderboard


def _bold_row(sheet, row_index: int) -> None:
    for cell in sheet[row_index]:
        cell.font = Font(bold=True)


def _write_task_analytics_sheet(wb, analytics, task_name_map, title):
    sheet = wb.create_sheet(title)
    sheet.append(
        ["Task", "Type", "Submissions", "Completion %", "Yes count", "Total", "Average", "Maximum", "Minimum", "Median"]
    )
    _bold_row(sheet, 1)
    for row in analytics.tasks:
        sheet.append(
            [
                row.task_name,
                row.task_type.value,
                row.submissions,
                round(row.completion_rate * 100, 1),
                row.yes_count if row.yes_count is not None else "",
                row.total if row.total is not None else "",
                round(row.average, 2) if row.average is not None else "",
                row.maximum if row.maximum is not None else "",
                row.minimum if row.minimum is not None else "",
                row.median if row.median is not None else "",
            ]
        )
    return sheet


def _write_student_analytics_sheet(wb, analytics, task_name_map, title):
    sheet = wb.create_sheet(title)
    breakdown_task_names = [task_name_map.get(task_id, str(task_id)) for task_id in analytics.breakdown_task_ids]
    sheet.append(
        ["Student", "Days participated", "Longest streak", "Best entry", *breakdown_task_names]
    )
    _bold_row(sheet, 1)
    for row in analytics.students:
        totals = {item.task_id: item.total for item in row.per_task_totals}
        sheet.append(
            [
                row.student_name,
                row.days_participated,
                row.longest_streak,
                row.best_entry if row.best_entry is not None else "",
                *[totals.get(task_id, 0) for task_id in analytics.breakdown_task_ids],
            ]
        )
    return sheet


def _write_daily_participation_sheet(wb, analytics, title):
    sheet = wb.create_sheet(title)
    sheet.append(["Date", "Active students"])
    _bold_row(sheet, 1)
    for row in analytics.daily_participation:
        sheet.append([row.date.isoformat(), row.active_students])
    return sheet


def export_room_to_workbook(
    db: Session,
    room: Room,
    from_date: date | None = None,
    to_date: date | None = None,
    student_id: int | None = None,
    leaderboard_only: bool = False,
    progress_only: bool = False,
) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)

    students_query = select(Student).where(Student.room_id == room.id)
    if student_id:
        students_query = students_query.where(Student.id == student_id)
    students = db.scalars(students_query.order_by(Student.name.asc())).all()
    tasks = db.scalars(select(Task).where(Task.room_id == room.id).order_by(Task.sort_order.asc(), Task.id.asc())).all()
    visible_tasks = [task for task in tasks if task.name != BONUS_TASK_NAME]
    leaderboard = build_leaderboard(db, room, respect_visibility=False)

    if not leaderboard_only and not progress_only:
        sheet = workbook.create_sheet("Room Info")
        sheet.append(["Field", "Value"])
        _bold_row(sheet, 1)
        sheet.append(["Room name", room.name])
        sheet.append(["Room code", room.room_code])
        sheet.append(["Status", room.status.value])
        sheet.append(["Participants", len(students)])
        sheet.append(["Tasks", len(visible_tasks)])
        sheet.append(["Created at", room.created_at.isoformat()])

        participants_sheet = workbook.create_sheet("Students")
        participants_sheet.append(
            [
                "ID",
                "Name",
                "Alias",
                "Telegram username",
                "Telegram ID",
                "Status",
                "Registered at",
                "Last submission",
                "Total score",
            ]
        )
        _bold_row(participants_sheet, 1)
        leaderboard_map = {entry.student_id: entry.total_points for entry in leaderboard}
        for student in students:
            participants_sheet.append(
                [
                    student.id,
                    student.name,
                    student.alias,
                    student.telegram_username,
                    student.telegram_id,
                    student.status.value,
                    student.created_at.isoformat(),
                    student.last_submission_at.isoformat() if student.last_submission_at else "",
                    leaderboard_map.get(student.id, 0),
                ]
            )

        tasks_sheet = workbook.create_sheet("Tasks")
        tasks_sheet.append(["ID", "Name", "Type", "Target", "Max", "Points", "Bonus", "Required", "Active"])
        _bold_row(tasks_sheet, 1)
        for task in visible_tasks:
            tasks_sheet.append(
                [
                    task.id,
                    task.name,
                    task.type.value,
                    float(task.target) if task.target is not None else "",
                    float(task.target_max) if task.target_max is not None else "",
                    task.points,
                    task.bonus_per_unit,
                    "Yes" if task.is_required else "No",
                    "Yes" if task.is_active else "No",
                ]
            )

    completions_query = select(Completion).where(Completion.room_id == room.id)
    if from_date:
        completions_query = completions_query.where(Completion.date >= from_date)
    if to_date:
        completions_query = completions_query.where(Completion.date <= to_date)
    if student_id:
        completions_query = completions_query.where(Completion.student_id == student_id)
    completions = db.scalars(completions_query.order_by(Completion.date.asc())).all()

    progress_sheet = workbook.create_sheet("Daily Results")
    header = ["Date", "Student"] + [task.name for task in visible_tasks] + ["Day points", "Total score"]
    progress_sheet.append(header)
    _bold_row(progress_sheet, 1)

    grouped = defaultdict(list)
    for completion in completions:
        grouped[(completion.date, completion.student_id)].append(completion)

    total_rows = db.execute(
        select(Completion.student_id, func.coalesce(func.sum(Completion.points_earned), 0)).where(Completion.room_id == room.id).group_by(Completion.student_id)
    ).all()
    total_map = {student_id: float(total) for student_id, total in total_rows}
    total_bonus_by_student, _ = build_bonus_maps(db, room)
    total_map = {student_id: total_map.get(student_id, 0) + total_bonus_by_student.get(student_id, 0) for student_id in set(total_map) | set(total_bonus_by_student)}

    student_map = {student.id: student for student in students}
    for (completion_date, completion_student_id), items in grouped.items():
        task_values = {item.task_id: item.value for item in items}
        _, day_bonus_map = build_bonus_maps(db, room, today=completion_date)
        day_points = sum(item.points_earned for item in items) + day_bonus_map.get(completion_student_id, 0)
        student = student_map.get(completion_student_id)
        if not student:
            continue
        progress_sheet.append(
            [
                completion_date.isoformat(),
                student.name,
                *[task_values.get(task.id, "") for task in visible_tasks],
                day_points,
                total_map.get(completion_student_id, 0),
            ]
        )

    leaderboard_sheet = workbook.create_sheet("Leaderboard")
    leaderboard_sheet.append(
        ["Place", "Participant", "Total score", "Today score", "Completed days", "Last submission"]
    )
    _bold_row(leaderboard_sheet, 1)
    for row in leaderboard:
        leaderboard_sheet.append(
            [
                row.position,
                row.display_name,
                row.total_points,
                row.today_points,
                row.completed_days,
                row.last_submission_at.isoformat() if row.last_submission_at else "",
            ]
        )

    if not leaderboard_only:
        punishments_sheet = workbook.create_sheet("Punishments")
        punishments_sheet.append(["Date", "Participant", "Type", "Reason", "Status"])
        _bold_row(punishments_sheet, 1)
        punishments = db.scalars(select(Punishment).where(Punishment.room_id == room.id).order_by(Punishment.assigned_at.desc())).all()
        for punishment in punishments:
            punishments_sheet.append(
                [
                    punishment.assigned_at.date().isoformat(),
                    punishment.student.name,
                    punishment.type,
                    punishment.reason,
                    punishment.status.value,
                ]
            )

    if not leaderboard_only and not progress_only:
        analytics = build_room_analytics(db, room)
        task_name_map = {task.id: task.name for task in visible_tasks}

        _write_task_analytics_sheet(workbook, analytics, task_name_map, "Task Analytics")
        _write_student_analytics_sheet(workbook, analytics, task_name_map, "Student Analytics")
        _write_daily_participation_sheet(workbook, analytics, "Daily Participation")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_multi_room_to_workbook(db: Session, rooms: list[Room]) -> BytesIO:
    workbook = Workbook()
    workbook.remove(workbook.active)

    room_analytics = [(room, build_room_analytics(db, room)) for room in rooms]
    all_days: set[date] = set()
    total_students = 0
    for room, analytics in room_analytics:
        total_students += len(room.students)
        for dp in analytics.daily_participation:
            all_days.add(dp.date)

    overview_sheet = workbook.create_sheet("Overview")
    overview_sheet.append(["Room count", len(rooms)])
    overview_sheet.append(["Total students", total_students])
    overview_sheet.append(["Total distinct days", len(all_days)])
    overview_sheet.append([])
    overview_sheet.append(["Room", "Students", "Distinct days"])
    _bold_row(overview_sheet, 5)
    for room, analytics in room_analytics:
        overview_sheet.append([room.name, len(room.students), len(analytics.daily_participation)])

    agg = build_multi_room_analytics(db, rooms)

    leaderboard_sheet = workbook.create_sheet("Combined Leaderboard")
    leaderboard_sheet.append(
        ["Rank", "Student", "Room", "Total points", "Today points", "Completed days"]
    )
    _bold_row(leaderboard_sheet, 1)
    for entry in agg.combined_leaderboard:
        leaderboard_sheet.append(
            [
                entry.position,
                entry.student_name,
                entry.room_name,
                entry.total_points,
                entry.today_points,
                entry.completed_days,
            ]
        )

    comparison_sheet = workbook.create_sheet("Room Comparison")
    comparison_sheet.append(["Room", "Students", "Total points", "Average points"])
    _bold_row(comparison_sheet, 1)
    for entry in agg.room_comparison:
        comparison_sheet.append(
            [
                entry.room_name,
                entry.student_count,
                entry.total_points,
                round(entry.average_points, 1),
            ]
        )

    for room, analytics in room_analytics:
        tasks = db.scalars(
            select(Task).where(Task.room_id == room.id).order_by(Task.sort_order.asc(), Task.id.asc())
        ).all()
        visible_tasks = [task for task in tasks if task.name != BONUS_TASK_NAME]
        task_name_map = {task.id: task.name for task in visible_tasks}

        _write_task_analytics_sheet(workbook, analytics, task_name_map, f"R{room.id} Tasks"[:31])
        _write_student_analytics_sheet(workbook, analytics, task_name_map, f"R{room.id} Students"[:31])
        _write_daily_participation_sheet(workbook, analytics, f"R{room.id} Daily"[:31])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
